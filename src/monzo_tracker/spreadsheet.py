"""Excel spreadsheet writer for Monzo transactions."""

import os
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import FormattedTransaction

# File path for the spreadsheet
SPREADSHEET_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "spreadsheet.xlsx")

# Styling constants
HEADER_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")
CATEGORY_FONT = Font(bold=True, size=10, italic=True)
COL_HEADER_FONT = Font(bold=True, size=10)
SUMMARY_FONT = Font(bold=True, size=11)
BALANCE_FONT = Font(bold=True, size=12)

OUT_FILL = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
IN_FILL = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
CATEGORY_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
BALANCE_FILL = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")

# Projected (future month) styling
PROJECTED_FONT = Font(italic=True, color="999999")
PROJECTED_SUMMARY_FONT = Font(bold=True, italic=True, size=11, color="999999")
PROJECTED_BALANCE_FONT = Font(bold=True, italic=True, size=12, color="999999")
PROJECTED_CAT_TOTAL_FONT = Font(bold=True, italic=True, size=9, color="999999")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)
THICK_BORDER = Border(
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)

# Column layout: A=Date, B=Description, C=Amount, D=ID (hidden for dedup)
COL_DATE = 1
COL_DESC = 2
COL_AMOUNT = 3
COL_ID = 4

COLUMN_WIDTHS = {COL_DATE: 20, COL_DESC: 35, COL_AMOUNT: 15, COL_ID: 5}


def _month_key(tx: FormattedTransaction) -> str:
    """Get month key like '2026-01' from a transaction."""
    return tx.date[:7]


def _month_sheet_name(month_key: str) -> str:
    """Convert '2026-01' to 'January 2026'."""
    dt = datetime.strptime(month_key, "%Y-%m")
    return dt.strftime("%B %Y")


def _get_existing_ids(ws) -> set[str]:
    """Read all transaction IDs already in a worksheet (column D)."""
    ids = set()
    for row in ws.iter_rows(min_col=COL_ID, max_col=COL_ID, values_only=False):
        cell = row[0]
        if cell.value and str(cell.value).startswith("tx_"):
            ids.add(str(cell.value))
    return ids


def _get_previous_balance(wb, current_month_key: str) -> float:
    """Get the running balance from the previous month's sheet."""
    # Collect all month sheet names and sort them
    month_sheets = []
    for name in wb.sheetnames:
        try:
            dt = datetime.strptime(name, "%B %Y")
            mk = dt.strftime("%Y-%m")
            month_sheets.append((mk, name))
        except ValueError:
            continue

    month_sheets.sort(key=lambda x: x[0])

    # Find the sheet just before the current month
    prev_name = None
    for mk, name in month_sheets:
        if mk >= current_month_key:
            break
        prev_name = name

    if prev_name is None:
        return 0.0

    # Read the running balance from the previous sheet (look for it from bottom up)
    ws = wb[prev_name]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=COL_DATE, max_col=COL_AMOUNT):
        if row[0].value == "Running Balance":
            return float(row[2].value or 0)
    return 0.0


def write_transactions(
    transactions: list[FormattedTransaction],
    filepath: str | None = None,
) -> str:
    """Write transactions to an Excel spreadsheet, grouped by month.

    Transactions are deduplicated by ID. Each month gets its own sheet
    with Out/In sections grouped by category.

    Args:
        transactions: List of formatted transactions to write.
        filepath: Path to the spreadsheet file. Defaults to spreadsheet.xlsx
                  in the project root.

    Returns:
        The filepath written to.
    """
    if filepath is None:
        filepath = os.path.abspath(SPREADSHEET_PATH)

    # Group transactions by month
    by_month: dict[str, list[FormattedTransaction]] = defaultdict(list)
    for tx in transactions:
        by_month[_month_key(tx)].append(tx)

    # Load or create workbook
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        # Remove the default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Process each month
    sorted_months = sorted(by_month.keys())
    for month_key in sorted_months:
        sheet_name = _month_sheet_name(month_key)
        month_txs = by_month[month_key]

        if sheet_name in wb.sheetnames:
            # Sheet exists — get existing IDs and filter out duplicates
            ws = wb[sheet_name]
            existing_ids = _get_existing_ids(ws)
            new_txs = [tx for tx in month_txs if tx.id not in existing_ids]
            if not new_txs:
                continue
            # Rebuild the sheet with merged data: existing + new
            all_txs = _read_transactions_from_sheet(ws) + new_txs
            del wb[sheet_name]
        else:
            all_txs = month_txs

        ws = wb.create_sheet(title=sheet_name)
        _write_month_sheet(ws, sheet_name, all_txs)

    # Sort sheets chronologically
    _sort_sheets(wb)

    # Recalculate running balances across all sheets
    _recalculate_balances(wb)

    # Build yearly overview sheets
    _write_yearly_overviews(wb)

    wb.save(filepath)
    return filepath


def _read_transactions_from_sheet(ws) -> list[FormattedTransaction]:
    """Read existing transactions from a worksheet to preserve them on rebuild."""
    transactions = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=COL_ID):
        id_cell = row[COL_ID - 1]
        if not id_cell.value or not str(id_cell.value).startswith("tx_"):
            continue

        date_cell = row[COL_DATE - 1]
        desc_cell = row[COL_DESC - 1]
        amount_cell = row[COL_AMOUNT - 1]

        # Determine category from the nearest category header above this row
        category = _find_category_for_row(ws, id_cell.row)

        # Determine if it's income or expense from the section
        section = _find_section_for_row(ws, id_cell.row)
        amount_val = float(amount_cell.value or 0)
        if section == "OUT":
            amount_raw = -abs(amount_val)
        else:
            amount_raw = abs(amount_val)

        notes = ""
        if desc_cell.comment:
            notes = desc_cell.comment.text

        transactions.append(FormattedTransaction(
            id=str(id_cell.value),
            date=str(date_cell.value),
            description=str(desc_cell.value),
            amount=f"GBP {amount_raw:.2f}",
            amount_raw=amount_raw,
            currency="GBP",
            category=category,
            notes=notes,
        ))
    return transactions


def _find_category_for_row(ws, target_row: int) -> str:
    """Walk upward from target_row to find the nearest category header."""
    for r in range(target_row - 1, 0, -1):
        cell = ws.cell(row=r, column=COL_DATE)
        # Category rows have the category fill and italic font
        if cell.font and cell.font.italic and cell.font.bold:
            return str(cell.value or "unknown")
        # Stop if we hit a section header
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb in ("FFD32F2F", "FF388E3C"):
            break
    return "unknown"


def _find_section_for_row(ws, target_row: int) -> str:
    """Walk upward from target_row to find if we're in OUT or IN section."""
    for r in range(target_row - 1, 0, -1):
        cell = ws.cell(row=r, column=COL_DATE)
        val = str(cell.value or "")
        if val == "OUT":
            return "OUT"
        if val == "IN":
            return "IN"
    return "OUT"


def _write_month_sheet(ws, title: str, transactions: list[FormattedTransaction]):
    """Write a complete month sheet with Out/In sections."""
    # Set column widths
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    # Hide the ID column
    ws.column_dimensions[get_column_letter(COL_ID)].hidden = True

    # Separate into out/in
    out_txs = [tx for tx in transactions if tx.amount_raw < 0]
    in_txs = [tx for tx in transactions if tx.amount_raw > 0]
    # Skip zero-amount transactions (active card checks etc.)
    zero_txs = [tx for tx in transactions if tx.amount_raw == 0]
    # Include zero amounts in the out section for visibility
    out_txs.extend(zero_txs)

    row = 1

    # Title
    ws.cell(row=row, column=COL_DATE, value=title).font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=COL_DATE, end_row=row, end_column=COL_AMOUNT)
    row += 2

    # === OUT SECTION ===
    row = _write_section(ws, row, "OUT", out_txs, OUT_FILL, is_expense=True)

    row += 1

    # === IN SECTION ===
    row = _write_section(ws, row, "IN", in_txs, IN_FILL, is_expense=False)

    row += 1

    # === SUMMARY ===
    total_out = sum(abs(tx.amount_raw) for tx in out_txs)
    total_in = sum(abs(tx.amount_raw) for tx in in_txs)
    net_change = total_in - total_out

    # Net Change row
    for col in range(COL_DATE, COL_AMOUNT + 1):
        ws.cell(row=row, column=col).fill = SUMMARY_FILL
        ws.cell(row=row, column=col).border = THICK_BORDER
    ws.cell(row=row, column=COL_DATE, value="Net Change").font = SUMMARY_FONT
    change_cell = ws.cell(row=row, column=COL_AMOUNT, value=net_change)
    change_cell.font = SUMMARY_FONT
    change_cell.number_format = '#,##0.00'
    change_cell.alignment = Alignment(horizontal="right")
    row += 1

    # Running Balance row (will be recalculated later)
    for col in range(COL_DATE, COL_AMOUNT + 1):
        ws.cell(row=row, column=col).fill = BALANCE_FILL
        ws.cell(row=row, column=col).border = THICK_BORDER
    ws.cell(row=row, column=COL_DATE, value="Running Balance").font = BALANCE_FONT
    balance_cell = ws.cell(row=row, column=COL_AMOUNT, value=0.0)
    balance_cell.font = BALANCE_FONT
    balance_cell.number_format = '#,##0.00'
    balance_cell.alignment = Alignment(horizontal="right")


def _write_section(
    ws,
    start_row: int,
    section_name: str,
    transactions: list[FormattedTransaction],
    header_fill: PatternFill,
    is_expense: bool,
) -> int:
    """Write an Out or In section. Returns the next available row."""
    row = start_row

    # Section header
    for col in range(COL_DATE, COL_AMOUNT + 1):
        ws.cell(row=row, column=col).fill = header_fill
    ws.cell(row=row, column=COL_DATE, value=section_name).font = SECTION_FONT
    row += 1

    # Column headers
    ws.cell(row=row, column=COL_DATE, value="Date").font = COL_HEADER_FONT
    ws.cell(row=row, column=COL_DESC, value="Description").font = COL_HEADER_FONT
    amount_header = ws.cell(row=row, column=COL_AMOUNT, value="Amount")
    amount_header.font = COL_HEADER_FONT
    amount_header.alignment = Alignment(horizontal="right")
    for col in range(COL_DATE, COL_AMOUNT + 1):
        ws.cell(row=row, column=col).border = Border(bottom=Side(style="medium"))
    row += 1

    # Group by category
    by_category: dict[str, list[FormattedTransaction]] = defaultdict(list)
    for tx in transactions:
        by_category[tx.category].append(tx)

    section_total = 0.0

    for category in sorted(by_category.keys()):
        cat_txs = by_category[category]
        # Sort by date within category
        cat_txs.sort(key=lambda x: x.date)

        # Category header
        for col in range(COL_DATE, COL_AMOUNT + 1):
            ws.cell(row=row, column=col).fill = CATEGORY_FILL
        ws.cell(row=row, column=COL_DATE, value=category).font = CATEGORY_FONT
        row += 1

        cat_total = 0.0
        for tx in cat_txs:
            display_amount = abs(tx.amount_raw)
            cat_total += display_amount

            ws.cell(row=row, column=COL_DATE, value=tx.date)
            desc_cell = ws.cell(row=row, column=COL_DESC, value=tx.description)
            amount_cell = ws.cell(row=row, column=COL_AMOUNT, value=display_amount)
            amount_cell.number_format = '#,##0.00'
            amount_cell.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=COL_ID, value=tx.id)

            # Add notes as a comment on the description cell
            if tx.notes:
                desc_cell.comment = Comment(tx.notes, "Monzo Tracker")

            for col in range(COL_DATE, COL_AMOUNT + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER

            row += 1

        # Category subtotal
        for col in range(COL_DATE, COL_AMOUNT + 1):
            ws.cell(row=row, column=col).fill = SUBTOTAL_FILL
        ws.cell(row=row, column=COL_DESC, value=f"{category} subtotal").font = Font(italic=True, size=9)
        subtotal_cell = ws.cell(row=row, column=COL_AMOUNT, value=cat_total)
        subtotal_cell.font = Font(italic=True, size=9)
        subtotal_cell.number_format = '#,##0.00'
        subtotal_cell.alignment = Alignment(horizontal="right")
        for col in range(COL_DATE, COL_AMOUNT + 1):
            ws.cell(row=row, column=col).border = Border(top=Side(style="thin"))
        row += 1

        section_total += cat_total

    # Section total
    for col in range(COL_DATE, COL_AMOUNT + 1):
        ws.cell(row=row, column=col).border = THICK_BORDER
    ws.cell(row=row, column=COL_DATE, value=f"TOTAL {section_name}").font = SUMMARY_FONT
    total_cell = ws.cell(row=row, column=COL_AMOUNT, value=section_total)
    total_cell.font = SUMMARY_FONT
    total_cell.number_format = '#,##0.00'
    total_cell.alignment = Alignment(horizontal="right")
    row += 1

    return row


def _sort_sheets(wb):
    """Sort workbook sheets: yearly overviews first, then monthly chronologically."""
    sheet_order = []
    for name in wb.sheetnames:
        # Yearly overview sheets like "2026 Overview"
        if name.endswith(" Overview"):
            try:
                year = int(name.split(" ")[0])
                # Sort key: Jan 1 of that year, but slightly before monthly sheets
                sheet_order.append((datetime(year, 1, 1, 0, 0, 0), name))
            except (ValueError, IndexError):
                sheet_order.append((datetime.max, name))
            continue

        # Monthly sheets like "January 2026"
        try:
            dt = datetime.strptime(name, "%B %Y")
            # Offset by 1 hour so monthly sheets sort after the yearly overview
            sheet_order.append((datetime(dt.year, dt.month, 1, 1, 0, 0), name))
        except ValueError:
            sheet_order.append((datetime.max, name))

    sheet_order.sort(key=lambda x: x[0])
    wb._sheets = [wb[name] for _, name in sheet_order]


def _recalculate_balances(wb):
    """Recalculate running balances across all month sheets in order."""
    balance = 0.0
    for ws in wb._sheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=COL_DATE, max_col=COL_AMOUNT):
            if row[0].value == "Net Change":
                net_change = float(row[COL_AMOUNT - 1].value or 0)
            if row[0].value == "Running Balance":
                balance += net_change
                row[COL_AMOUNT - 1].value = balance


def _extract_month_summary(ws) -> dict:
    """Extract category totals and summary values from a monthly sheet.

    Returns a dict with:
        out_categories: {category: total}
        in_categories: {category: total}
        total_out: float
        total_in: float
        net_change: float
        running_balance: float
    """
    result = {
        "out_categories": {},
        "in_categories": {},
        "total_out": 0.0,
        "total_in": 0.0,
        "net_change": 0.0,
        "running_balance": 0.0,
    }

    current_section = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3, values_only=False):
        label = row[0].value
        amount = row[2].value

        if label == "OUT":
            current_section = "out"
            continue
        if label == "IN":
            current_section = "in"
            continue

        # Category subtotal rows
        desc = row[1].value
        if desc and isinstance(desc, str) and desc.endswith(" subtotal"):
            cat_name = desc.replace(" subtotal", "")
            cat_total = float(amount or 0)
            if current_section == "out":
                result["out_categories"][cat_name] = cat_total
            elif current_section == "in":
                result["in_categories"][cat_name] = cat_total
            continue

        if label == "TOTAL OUT":
            result["total_out"] = float(amount or 0)
        elif label == "TOTAL IN":
            result["total_in"] = float(amount or 0)
        elif label == "Net Change":
            result["net_change"] = float(amount or 0)
        elif label == "Running Balance":
            result["running_balance"] = float(amount or 0)

    return result


def _write_yearly_overviews(wb):
    """Create/rebuild yearly overview sheets from monthly data."""
    # Collect monthly sheets grouped by year
    by_year: dict[str, list[tuple[str, str]]] = defaultdict(list)  # year -> [(month_key, sheet_name)]
    for name in wb.sheetnames:
        try:
            dt = datetime.strptime(name, "%B %Y")
            year = str(dt.year)
            month_key = dt.strftime("%Y-%m")
            by_year[year].append((month_key, name))
        except ValueError:
            continue

    for year, months in by_year.items():
        months.sort(key=lambda x: x[0])
        overview_name = f"{year} Overview"

        # Remove existing overview sheet to rebuild
        if overview_name in wb.sheetnames:
            del wb[overview_name]

        ws = wb.create_sheet(title=overview_name)
        _write_yearly_sheet(wb, ws, year, months)

    # Re-sort so overviews appear before their year's months
    _sort_sheets(wb)


YEAR_LABEL_WIDTH = 22
YEAR_MONTH_WIDTH = 14


def _write_yearly_sheet(wb, ws, year: str, months: list[tuple[str, str]]):
    """Write a yearly overview sheet with all 12 months as columns.

    Months with actual data show real values. Future months are pre-filled
    with Excel AVERAGE formulas based on existing data, and running balance
    is extrapolated forward.
    """
    # All 12 months for the year
    all_month_keys = [f"{year}-{m:02d}" for m in range(1, 13)]
    month_short_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                         "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    # Which months have actual data
    data_month_set = set(mk for mk, _ in months)
    num_data = len(months)

    # Extract summaries from each monthly sheet
    summaries = {}
    all_out_cats = set()
    all_in_cats = set()
    for month_key, sheet_name in months:
        summary = _extract_month_summary(wb[sheet_name])
        summaries[month_key] = summary
        all_out_cats.update(summary["out_categories"].keys())
        all_in_cats.update(summary["in_categories"].keys())

    out_cats = sorted(all_out_cats)
    in_cats = sorted(all_in_cats)

    # Columns: A=labels, B=Jan, C=Feb, ..., M=Dec, N=Total
    total_col = 14

    # Column indices (2-based) for months with actual data
    data_col_indices = []
    for i, mk in enumerate(all_month_keys):
        if mk in data_month_set:
            data_col_indices.append(i + 2)

    # Formula helpers
    def avg_of_data(r):
        """SUM of data month cells / count — treats empty cells as 0."""
        refs = ",".join(f"{get_column_letter(c)}{r}" for c in data_col_indices)
        return f"=SUM({refs})/{num_data}"

    def sum_row(r):
        """SUM across all 12 month columns for the Total column."""
        return f"=SUM({get_column_letter(2)}{r}:{get_column_letter(13)}{r})"

    def sum_cells(col, rows):
        """SUM of specific cells in a column (for TOTAL rows from categories)."""
        cl = get_column_letter(col)
        refs = ",".join(f"{cl}{r}" for r in rows)
        return f"=SUM({refs})"

    # Column widths
    ws.column_dimensions["A"].width = YEAR_LABEL_WIDTH
    for i in range(12):
        ws.column_dimensions[get_column_letter(i + 2)].width = YEAR_MONTH_WIDTH
    ws.column_dimensions[get_column_letter(total_col)].width = YEAR_MONTH_WIDTH

    row = 1

    # Title
    ws.cell(row=row, column=1, value=f"{year} Overview").font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min(total_col, 6))
    row += 2

    # Month headers
    for i, name in enumerate(month_short_names):
        col = i + 2
        cell = ws.cell(row=row, column=col, value=name)
        cell.font = COL_HEADER_FONT
        cell.alignment = Alignment(horizontal="right")
        cell.border = Border(bottom=Side(style="medium"))
    total_header = ws.cell(row=row, column=total_col, value="Total")
    total_header.font = Font(bold=True, size=10)
    total_header.alignment = Alignment(horizontal="right")
    total_header.border = Border(bottom=Side(style="medium"))
    ws.cell(row=row, column=1).border = Border(bottom=Side(style="medium"))
    row += 1

    # === OUT SECTION ===
    for col in range(1, total_col + 1):
        ws.cell(row=row, column=col).fill = OUT_FILL
    ws.cell(row=row, column=1, value="OUT").font = SECTION_FONT
    row += 1

    out_cat_rows = []
    for cat in out_cats:
        ws.cell(row=row, column=1, value=f"  {cat}").font = CATEGORY_FONT
        out_cat_rows.append(row)

        for i, mk in enumerate(all_month_keys):
            col = i + 2
            if mk in data_month_set:
                val = summaries[mk]["out_categories"].get(cat, 0.0)
                if val:
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            else:
                cell = ws.cell(row=row, column=col, value=avg_of_data(row))
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                cell.font = PROJECTED_FONT

        tc = ws.cell(row=row, column=total_col, value=sum_row(row))
        tc.number_format = '#,##0.00'
        tc.alignment = Alignment(horizontal="right")
        tc.font = Font(bold=True, size=9)
        for col in range(1, total_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # TOTAL OUT
    total_out_row = row
    for col in range(1, total_col + 1):
        ws.cell(row=row, column=col).border = THICK_BORDER
    ws.cell(row=row, column=1, value="TOTAL OUT").font = SUMMARY_FONT
    for i, mk in enumerate(all_month_keys):
        col = i + 2
        if mk in data_month_set:
            val = summaries[mk]["total_out"]
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = SUMMARY_FONT
        else:
            cell = ws.cell(row=row, column=col, value=sum_cells(col, out_cat_rows))
            cell.font = PROJECTED_SUMMARY_FONT
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
    tc = ws.cell(row=row, column=total_col, value=sum_row(row))
    tc.number_format = '#,##0.00'
    tc.alignment = Alignment(horizontal="right")
    tc.font = SUMMARY_FONT
    row += 2

    # === IN SECTION ===
    for col in range(1, total_col + 1):
        ws.cell(row=row, column=col).fill = IN_FILL
    ws.cell(row=row, column=1, value="IN").font = SECTION_FONT
    row += 1

    in_cat_rows = []
    for cat in in_cats:
        ws.cell(row=row, column=1, value=f"  {cat}").font = CATEGORY_FONT
        in_cat_rows.append(row)

        for i, mk in enumerate(all_month_keys):
            col = i + 2
            if mk in data_month_set:
                val = summaries[mk]["in_categories"].get(cat, 0.0)
                if val:
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            else:
                cell = ws.cell(row=row, column=col, value=avg_of_data(row))
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                cell.font = PROJECTED_FONT

        tc = ws.cell(row=row, column=total_col, value=sum_row(row))
        tc.number_format = '#,##0.00'
        tc.alignment = Alignment(horizontal="right")
        tc.font = Font(bold=True, size=9)
        for col in range(1, total_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # TOTAL IN
    total_in_row = row
    for col in range(1, total_col + 1):
        ws.cell(row=row, column=col).border = THICK_BORDER
    ws.cell(row=row, column=1, value="TOTAL IN").font = SUMMARY_FONT
    for i, mk in enumerate(all_month_keys):
        col = i + 2
        if mk in data_month_set:
            val = summaries[mk]["total_in"]
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = SUMMARY_FONT
        else:
            cell = ws.cell(row=row, column=col, value=sum_cells(col, in_cat_rows))
            cell.font = PROJECTED_SUMMARY_FONT
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
    tc = ws.cell(row=row, column=total_col, value=sum_row(row))
    tc.number_format = '#,##0.00'
    tc.alignment = Alignment(horizontal="right")
    tc.font = SUMMARY_FONT
    row += 2

    # === NET CHANGE ===
    net_change_row = row
    for col in range(1, total_col + 1):
        ws.cell(row=row, column=col).fill = SUMMARY_FILL
        ws.cell(row=row, column=col).border = THICK_BORDER
    ws.cell(row=row, column=1, value="Net Change").font = SUMMARY_FONT
    for i, mk in enumerate(all_month_keys):
        col = i + 2
        cl = get_column_letter(col)
        if mk in data_month_set:
            val = summaries[mk]["net_change"]
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = SUMMARY_FONT
        else:
            cell = ws.cell(row=row, column=col,
                           value=f"={cl}{total_in_row}-{cl}{total_out_row}")
            cell.font = PROJECTED_SUMMARY_FONT
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
    tc = ws.cell(row=row, column=total_col, value=sum_row(row))
    tc.number_format = '#,##0.00'
    tc.alignment = Alignment(horizontal="right")
    tc.font = SUMMARY_FONT
    row += 1

    # === RUNNING BALANCE ===
    balance_row = row
    for col in range(1, total_col + 1):
        ws.cell(row=row, column=col).fill = BALANCE_FILL
        ws.cell(row=row, column=col).border = THICK_BORDER
    ws.cell(row=row, column=1, value="Running Balance").font = BALANCE_FONT
    for i, mk in enumerate(all_month_keys):
        col = i + 2
        cl = get_column_letter(col)
        if mk in data_month_set:
            val = summaries[mk]["running_balance"]
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = BALANCE_FONT
        else:
            if col == 2:
                # First month (Jan) with no data: balance = net change
                formula = f"={cl}{net_change_row}"
            else:
                prev_cl = get_column_letter(col - 1)
                formula = f"={prev_cl}{balance_row}+{cl}{net_change_row}"
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = PROJECTED_BALANCE_FONT
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
    # Total column: last month's running balance (Dec)
    tc = ws.cell(row=row, column=total_col,
                 value=f"={get_column_letter(13)}{balance_row}")
    tc.number_format = '#,##0.00'
    tc.alignment = Alignment(horizontal="right")
    tc.font = BALANCE_FONT
